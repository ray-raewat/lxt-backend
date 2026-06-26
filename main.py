from fastapi import FastAPI, File, UploadFile, Form, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import datetime, json, os, requests, bcrypt, jwt as pyjwt

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

SUPABASE_URL      = "https://woicdagfrkmtxhmqteyy.supabase.co"
SUPABASE_KEY      = os.environ.get("SUPABASE_KEY", "")
JWT_SECRET        = os.environ.get("JWT_SECRET", "lxt-dev-secret")
JWT_ALGO          = "HS256"
JWT_HOURS         = 168  # 7 days
MAKE_WEBHOOK_URL  = os.environ.get("MAKE_WEBHOOK_URL", "")   # Make "Custom Webhook" URL
LINE_NOTIFY_TOKEN = os.environ.get("LINE_NOTIFY_TOKEN", "")  # LINE Notify token ของ ectramu
BACKEND_URL       = os.environ.get("BACKEND_URL", "http://localhost:8000")

security = HTTPBearer(auto_error=False)

# ── Helpers ───────────────────────────────────────────────────────────────────

def sb_headers():
    return {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json", "Prefer": "return=representation"}

def notify_make(payload: dict):
    """ส่งข้อมูลไปที่ Make webhook → Make จะส่ง LINE message ต่อ"""
    if not MAKE_WEBHOOK_URL:
        return
    try:
        requests.post(MAKE_WEBHOOK_URL, json=payload, timeout=5)
    except Exception as e:
        print(f"⚠️ Make webhook error: {e}")

def notify_line(message: str):
    """ส่งข้อความตรงผ่าน LINE Notify (ถ้ามี token)"""
    if not LINE_NOTIFY_TOKEN:
        return
    try:
        requests.post(
            "https://notify-api.line.me/api/notify",
            headers={"Authorization": f"Bearer {LINE_NOTIFY_TOKEN}"},
            data={"message": message},
            timeout=5,
        )
    except Exception as e:
        print(f"⚠️ LINE Notify error: {e}")

def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def check_pw(pw: str, hashed: str) -> bool:
    return bcrypt.checkpw(pw.encode(), hashed.encode())

def make_token(uid, username, role, full_name):
    exp = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=JWT_HOURS)
    return pyjwt.encode({"sub": str(uid), "username": username, "role": role,
                         "full_name": full_name, "exp": exp}, JWT_SECRET, algorithm=JWT_ALGO)

def get_user(creds: HTTPAuthorizationCredentials = Depends(security)):
    if not creds:
        raise HTTPException(401, "Not authenticated")
    try:
        return pyjwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired – please login again")
    except Exception:
        raise HTTPException(401, "Invalid token")

def get_user_from_query(token: str = ""):
    """Accept token as query param (for window.open downloads)."""
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        return pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except Exception:
        raise HTTPException(401, "Invalid token")

def admin_only(user=Depends(get_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    return user

# ── Bootstrap: create default admin if no users exist ─────────────────────────

def bootstrap():
    r = requests.get(f"{SUPABASE_URL}/rest/v1/users?limit=1", headers=sb_headers())
    if not r.json():
        requests.post(f"{SUPABASE_URL}/rest/v1/users", headers=sb_headers(), json={
            "username": "admin", "password_hash": hash_pw("Admin@lxt2024"),
            "full_name": "Administrator", "role": "admin"
        })
        print("✅ Default admin created  (admin / Admin@lxt2024)")

try:
    bootstrap()
except Exception as e:
    print(f"⚠️ Bootstrap skipped: {e}")

# ── Auth ──────────────────────────────────────────────────────────────────────

@app.get("/")
def root(): return {"message": "LXT Backend Running"}

@app.get("/health")
def health():
    try:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?limit=1", headers=sb_headers(), timeout=10)
        return {"db": "connected", "status": r.status_code}
    except Exception as e:
        return {"db": "error", "detail": str(e)}

@app.get("/ping")
def ping():
    """Cron-job endpoint — queries DB to keep Supabase free tier active."""
    try:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?limit=1", headers=sb_headers(), timeout=10)
        return {"ok": True, "db": r.status_code, "ts": datetime.datetime.now().isoformat()}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/auth/login")
def login(data: dict):
    username = data.get("username", "").strip()
    password = data.get("password", "")
    r = requests.get(f"{SUPABASE_URL}/rest/v1/users?username=eq.{username}&active=eq.true", headers=sb_headers())
    rows = r.json()
    if not rows or not check_pw(password, rows[0]["password_hash"]):
        raise HTTPException(401, "Invalid username or password")
    u = rows[0]
    token = make_token(u["id"], u["username"], u["role"], u["full_name"])
    return {"token": token, "username": u["username"], "role": u["role"], "full_name": u["full_name"]}

@app.get("/auth/me")
def me(user=Depends(get_user)):
    return user

# ── User Management (admin only) ──────────────────────────────────────────────

@app.get("/users")
def list_users(admin=Depends(admin_only)):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/users?order=id.asc&select=id,username,full_name,role,active,created_at", headers=sb_headers())
    return r.json()

@app.post("/users")
def create_user(data: dict, admin=Depends(admin_only)):
    payload = {
        "username":      data.get("username", "").strip(),
        "password_hash": hash_pw(data.get("password", "changeme")),
        "full_name":     data.get("full_name", ""),
        "role":          data.get("role", "user"),
        "active":        True,
    }
    r = requests.post(f"{SUPABASE_URL}/rest/v1/users", headers=sb_headers(), json=payload)
    if r.status_code in (200, 201):
        return {"status": "created"}
    raise HTTPException(400, r.text)

@app.put("/users/{uid}")
def update_user(uid: int, data: dict, admin=Depends(admin_only)):
    payload = {k: v for k, v in {
        "username":  data.get("username"),
        "full_name": data.get("full_name"),
        "role":      data.get("role"),
        "active":    data.get("active"),
    }.items() if v is not None}
    if "password" in data and data["password"]:
        payload["password_hash"] = hash_pw(data["password"])
    r = requests.patch(f"{SUPABASE_URL}/rest/v1/users?id=eq.{uid}", headers=sb_headers(), json=payload)
    return {"status": "updated"}

@app.delete("/users/{uid}")
def delete_user(uid: int, admin=Depends(admin_only)):
    requests.delete(f"{SUPABASE_URL}/rest/v1/users?id=eq.{uid}", headers=sb_headers())
    return {"status": "deleted"}

# ── Image Upload ──────────────────────────────────────────────────────────────


@app.post("/upload-image")
async def upload_image(file: UploadFile = File(...), category: str = Form(...), user=Depends(get_user)):
    content = await file.read()
    ts = int(datetime.datetime.now().timestamp() * 1000)
    filename = f"{category}/{ts}.jpg"
    r = requests.post(f"{SUPABASE_URL}/storage/v1/object/report-images/{filename}",
        headers={"Authorization": f"Bearer {SUPABASE_KEY}", "Content-Type": "image/jpeg"}, data=content)
    if r.status_code in (200, 201):
        return {"url": f"{SUPABASE_URL}/storage/v1/object/public/report-images/{filename}"}
    err = r.text or f"HTTP {r.status_code}"
    print(f"⚠️ Upload failed {r.status_code}: {err[:200]}")
    return {"error": err}

# ── Reports CRUD ──────────────────────────────────────────────────────────────

def _parse(row):
    safety_raw = row.get("safety") or {}
    if isinstance(safety_raw, str):
        try: safety_raw = json.loads(safety_raw)
        except: safety_raw = {}
    return {
        "id":               row["id"],
        "date":             row["date"],
        "project":          row["project"],
        "site":             row["site"],
        "gps":              row["gps"],
        "submitted_by":     row.get("submitted_by", "") or "",
        "workTypes":        json.loads(row.get("work_types") or "[]"),
        "description":      row["description"],
        "quantity":         row["quantity"],
        "issues":           row["issues"],
        "teamImages":       json.loads(row.get("team_images") or "[]"),
        "equipmentImages":  json.loads(row.get("equipment_images") or "[]"),
        "materialImages":   _parse_captioned_list(row.get("material_images") or "[]"),
        "areaImages":       _parse_captioned_list(row.get("area_images") or "[]"),
        "closingImages":    json.loads(row.get("closing_images") or "[]"),
        "safety":           safety_raw,
    }

@app.post("/reports")
def create_report(data: dict, user=Depends(get_user)):
    payload = {
        "date":             data.get("date", datetime.date.today().isoformat()),
        "project":          data.get("project", ""),
        "site":             data.get("site", ""),
        "gps":              data.get("gps", ""),
        "work_types":       json.dumps(data.get("workTypes", [])),
        "description":      data.get("description", ""),
        "quantity":         data.get("quantity", ""),
        "issues":           data.get("issues", ""),
        "team_images":      json.dumps(data.get("teamImages", [])),
        "equipment_images": json.dumps(data.get("equipmentImages", [])),
        "area_images":      json.dumps(data.get("areaImages", [])),
        "material_images":  json.dumps(data.get("materialImages", [])),
        "closing_images":   json.dumps(data.get("closingImages", [])),
        "safety":           data.get("safety", {}),
        "submitted_by":     user.get("username", ""),
    }
    r = requests.post(f"{SUPABASE_URL}/rest/v1/reports", headers=sb_headers(), json=payload)
    if r.status_code in (200, 201):
        saved = r.json()
        rid = saved[0]["id"] if saved else None
        report_url = f"{BACKEND_URL}/report/{rid}" if rid else ""
        wt_list = data.get("workTypes", [])
        # แจ้งเตือนผ่าน Make → LINE
        notify_make({
            "event":       "new_report",
            "report_id":   rid,
            "date":        payload["date"],
            "project":     payload["project"],
            "site":        payload["site"],
            "work_types":  ", ".join(wt_list),
            "description": payload["description"],
            "quantity":    payload["quantity"],
            "issues":      payload["issues"],
            "report_url":  report_url,
            "submitted_by": user.get("full_name", user.get("username", "")),
        })
        # แจ้งเตือนผ่าน LINE Notify โดยตรง (backup)
        msg = (
            f"\n📋 รายงานใหม่ #{rid}"
            f"\n📅 {payload['date']}"
            f"\n🏗️ {payload['project']} | 📍 {payload['site']}"
            f"\n🔧 {', '.join(wt_list) or '-'}"
            f"\n📝 {payload['description'][:80]}"
            + (f"\n⚠️ {payload['issues']}" if payload.get("issues") else "")
            + (f"\n🔗 {report_url}" if report_url else "")
        )
        notify_line(msg)
        return {"status": "success", "report_id": rid}
    return {"status": "error", "detail": r.text}

@app.get("/reports")
def get_reports(user=Depends(get_user)):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?order=date.desc,id.desc", headers=sb_headers())
    return {"total": len(r.json()), "data": [_parse(row) for row in r.json()]}

@app.delete("/reports/{rid}")
def delete_report(rid: int, admin=Depends(admin_only)):
    requests.delete(f"{SUPABASE_URL}/rest/v1/reports?id=eq.{rid}", headers=sb_headers())
    return {"status": "deleted"}

@app.put("/reports/{rid}")
def update_report(rid: int, data: dict, user=Depends(get_user)):
    # Fetch existing report
    r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?id=eq.{rid}", headers=sb_headers())
    rows = r.json()
    if not rows:
        raise HTTPException(404, "Report not found")
    rep = rows[0]
    username = user.get("username", "")
    role_    = user.get("role", "")
    is_owner = rep.get("submitted_by", "") == username and username != ""
    is_admin = role_ == "admin"
    if not is_admin and not is_owner:
        raise HTTPException(403, "ไม่มีสิทธิ์แก้ไขรายงานนี้")
    today = datetime.date.today().isoformat()
    if not is_admin and rep.get("date", "") != today:
        raise HTTPException(403, "แก้ไขได้เฉพาะรายงานของวันนี้เท่านั้น")
    # Only allow these fields to be updated
    allowed = {"description", "quantity", "issues", "safety"}
    update = {k: v for k, v in data.items() if k in allowed}
    if not update:
        return {"status": "nothing_to_update"}
    resp = requests.patch(
        f"{SUPABASE_URL}/rest/v1/reports?id=eq.{rid}",
        headers=sb_headers(), json=update)
    if resp.status_code in (200, 201, 204):
        return {"status": "updated"}
    return {"status": "error", "detail": resp.text}

# ── Stats ─────────────────────────────────────────────────────────────────────

@app.get("/stats")
def get_stats(user=Depends(get_user)):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?order=date.asc", headers=sb_headers())
    rows = r.json()
    today = datetime.date.today().isoformat()
    week_start  = (datetime.date.today() - datetime.timedelta(days=6)).isoformat()
    month_start = datetime.date.today().replace(day=1).isoformat()
    wt = {}
    for row in rows:
        for t in json.loads(row.get("work_types") or "[]"):
            wt[t] = wt.get(t, 0) + 1
    thirty = (datetime.date.today() - datetime.timedelta(days=29)).isoformat()
    dm = {}
    for row in rows:
        d = row.get("date", "")
        if d >= thirty: dm[d] = dm.get(d, 0) + 1
    pm = {}
    for row in rows: p = row.get("project",""); pm[p] = pm.get(p,0)+1
    sm = {}
    for row in rows: s = row.get("site",""); sm[s] = sm.get(s,0)+1
    return {
        "total": len(rows),
        "today": sum(1 for r in rows if r.get("date") == today),
        "this_week":  sum(1 for r in rows if r.get("date","") >= week_start),
        "this_month": sum(1 for r in rows if r.get("date","") >= month_start),
        "by_work_type": [{"name":k,"count":v} for k,v in sorted(wt.items(),key=lambda x:-x[1])],
        "daily": [{"date":k,"count":v} for k,v in sorted(dm.items())],
        "by_project": sorted([{"project":k,"count":v} for k,v in pm.items()],key=lambda x:-x["count"])[:10],
        "by_site":    sorted([{"site":k,"count":v}    for k,v in sm.items()],key=lambda x:-x["count"])[:10],
    }

# ── LINE ↔ Make Webhook ───────────────────────────────────────────────────────

@app.post("/line/webhook")
async def line_webhook(data: dict):
    """
    รับข้อความจาก Make (Make ได้รับ LINE message แล้วส่งต่อมาที่นี่)
    Make ควรส่ง JSON: { "user_id": "...", "message": "...", "reply_token": "..." }
    """
    msg  = (data.get("message") or "").strip().lower()
    uid  = data.get("user_id", "")
    reply = ""

    if msg in ("สถิติ", "stat", "stats", "ยอด"):
        r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?order=date.asc", headers=sb_headers())
        rows = r.json()
        today = datetime.date.today().isoformat()
        wk    = (datetime.date.today() - datetime.timedelta(days=6)).isoformat()
        mo    = datetime.date.today().replace(day=1).isoformat()
        reply = (
            f"📊 สถิติ CoWork\n"
            f"ทั้งหมด: {len(rows)} รายงาน\n"
            f"วันนี้: {sum(1 for x in rows if x.get('date')==today)}\n"
            f"สัปดาห์นี้: {sum(1 for x in rows if x.get('date','')>=wk)}\n"
            f"เดือนนี้: {sum(1 for x in rows if x.get('date','')>=mo)}"
        )

    elif msg in ("รายงานวันนี้", "today", "วันนี้"):
        today = datetime.date.today().isoformat()
        r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?date=eq.{today}&order=id.desc", headers=sb_headers())
        rows = r.json()
        if not rows:
            reply = f"📋 วันนี้ ({today}) ยังไม่มีรายงาน"
        else:
            lines = [f"📋 รายงานวันนี้ {today} ({len(rows)} รายการ)"]
            for x in rows[:5]:
                lines.append(f"• #{x['id']} {x.get('project','')} – {x.get('site','')}")
            if len(rows) > 5:
                lines.append(f"...และอีก {len(rows)-5} รายการ")
            reply = "\n".join(lines)

    elif msg.startswith("รายงาน ") or msg.startswith("report "):
        parts = msg.split()
        if len(parts) >= 2 and parts[1].isdigit():
            rid = int(parts[1])
            reply = f"🔗 ดูรายงาน #{rid}: {BACKEND_URL}/report/{rid}"
        else:
            reply = "❌ รูปแบบ: รายงาน {id}  เช่น รายงาน 42"

    elif msg in ("help", "ช่วยเหลือ", "คำสั่ง", "?"):
        reply = (
            "📱 คำสั่ง CoWork\n"
            "• สถิติ – ดูสรุปยอด\n"
            "• วันนี้ – รายงานวันนี้\n"
            "• รายงาน {id} – ดูรายงานตาม ID\n"
            "• ช่วยเหลือ – คำสั่งทั้งหมด"
        )
    else:
        reply = f"❓ ไม่เข้าใจคำสั่ง \"{data.get('message','')}\" พิมพ์ 'ช่วยเหลือ' เพื่อดูคำสั่ง"

    return {"reply": reply, "user_id": uid}


# ── HTML Daily Report ─────────────────────────────────────────────────────────

def _grid(title, urls):
    if not urls: return f'<div class="section"><div class="sec-title">{title}</div><p class="empty">No photos uploaded</p></div>'
    imgs = "".join(f'<img src="{u}" onerror="this.style.display=\'none\'">' for u in urls)
    return f'<div class="section"><div class="sec-title">{title}</div><div class="photos">{imgs}</div></div>'

def _grid_cap(title, items):
    """Photo grid with captions — for Material and Work Area sections."""
    if not items: return f'<div class="section"><div class="sec-title">{title}</div><p class="empty">No photos uploaded</p></div>'
    inner = ""
    for item in items:
        url     = item.get("url","") if isinstance(item, dict) else item
        caption = item.get("caption","") if isinstance(item, dict) else ""
        cap_html = f'<p class="photo-caption">{caption}</p>' if caption else ""
        inner += f'<div class="photo-item"><img src="{url}" onerror="this.style.display:\'none\'">{cap_html}</div>'
    return f'<div class="section"><div class="sec-title">{title}</div><div class="photos-cap">{inner}</div></div>'

SAFETY_LABELS = [
    ("toolboxTalk",  "1.1 ประชุมความปลอดภัย ก่อนเริ่มงาน (Toolbox Talk)"),
    ("trafficSigns", "1.2 ตั้งกรวย แผงกั้น และป้ายเตือนครบถ้วน (Traffic Signs)"),
    ("shoring",      "1.3 ตรวจสอบความปลอดภัยหน้างานขุด (Shoring / Slope)"),
    ("ppe",          "1.4 พนักงานทุกคนสวม PPE ครบถ้วน (หมวก เสื้อกั๊ก รองเท้า)"),
    ("areaSafety",   "1.5 ตรวจสอบความปลอดภัยรอบพื้นที่ทำงาน (Area Safety)"),
]

# ── Report Display Settings (defaults — override via app_settings table) ───────
DEFAULT_DISPLAY = {
    "showGps":             True,
    "showWorkTypes":       True,
    "showQuantity":        True,
    "showIssues":          True,
    "showSafety":          True,
    "showLabor":           True,
    "showTeamImages":      True,
    "showEquipmentImages": True,
    "showMaterialImages":  True,
    "showAreaImages":      True,
    "showClosingImages":   True,
}

def _parse_captioned_list(raw):
    """Parse image list supporting both old (string) and new ({url,caption}) format."""
    items = json.loads(raw or "[]")
    result = []
    for item in items:
        if isinstance(item, str):
            result.append({"url": item, "caption": ""})
        elif isinstance(item, dict):
            result.append({"url": item.get("url", ""), "caption": item.get("caption", "")})
    return result

def _safety_section(safety, show_labor=True):
    if not safety: return ""
    rows_html = ""
    for key, label in SAFETY_LABELS:
        val = safety.get(key, False)
        icon = "✅" if val else "❌"
        bg   = "#f0fdf4" if val else "#fff"
        rows_html += f'<tr style="background:{bg}"><td style="padding:7px 12px;font-size:13px">{label}</td><td style="padding:7px 12px;text-align:center;font-size:16px">{icon}</td></tr>'
    if show_labor:
        labor = safety.get("laborCount", "-") or "-"
        rows_html += f'<tr style="background:#fffbeb"><td style="padding:7px 12px;font-size:13px">1.6 จำนวนคนงาน (Labor)</td><td style="padding:7px 12px;text-align:center;font-weight:700;font-size:14px">{labor} คน</td></tr>'
    checked = sum(1 for k,_ in SAFETY_LABELS if safety.get(k))
    total = len(SAFETY_LABELS)
    bar_color = "#16a34a" if checked==total else "#d97706" if checked>=total//2 else "#dc2626"
    return f'''<div class="section">
<div class="sec-title">🦺 Safety Checklist — {checked}/{total} รายการ</div>
<table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;border-radius:6px;overflow:hidden">
  <thead><tr style="background:#1e3a5f;color:white">
    <th style="padding:8px 12px;text-align:left;font-size:13px">รายการตรวจสอบ</th>
    <th style="padding:8px 12px;text-align:center;font-size:13px;width:80px">สถานะ</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div style="margin-top:8px;background:#f3f4f6;border-radius:4px;height:8px;overflow:hidden">
  <div style="width:{(checked/total)*100:.0f}%;height:100%;background:{bar_color};transition:width .3s"></div>
</div>
</div>'''

# ── Report Display Settings ───────────────────────────────────────────────────
# Strategy: in-memory cache (primary, always works within a process)
#            + Supabase reports table row id=-1 trick via upsert (persistence)
# We repurpose a single JSONB column by storing settings as a "virtual report"
# with id=0 in a dedicated key inside app state.  Simpler: just use a module-
# level dict as single source of truth and persist to Supabase via a dedicated
# text column we already have available on the users table as a fallback.
#
# Actual implementation: module-level dict + best-effort persist to Storage.
# The in-memory dict is authoritative; Storage is only used for cold-start load.

_ds_cache: dict = {}   # populated lazily on first read

def _storage_url_public() -> str:
    return f"{SUPABASE_URL}/storage/v1/object/public/report-images/system/display_settings.json"

def _storage_url_upload() -> str:
    return f"{SUPABASE_URL}/storage/v1/object/report-images/system/display_settings.json"

def _load_from_storage() -> dict:
    """Try to read persisted settings from Supabase Storage."""
    try:
        r = requests.get(_storage_url_public(), timeout=4)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, dict):
                print("✅ Settings loaded from Storage")
                return {**DEFAULT_DISPLAY, **data}
    except Exception as e:
        print(f"⚠️ Settings load: {e}")
    return DEFAULT_DISPLAY.copy()

def _persist_to_storage(value: dict):
    """Best-effort: save settings JSON to Supabase Storage."""
    try:
        hdrs = {"Authorization": f"Bearer {SUPABASE_KEY}",
                "Content-Type": "application/json",
                "x-upsert": "true"}
        r = requests.post(_storage_url_upload(), headers=hdrs,
                          data=json.dumps(value).encode(), timeout=6)
        if r.status_code in (200, 201):
            print("✅ Settings persisted to Storage")
        else:
            print(f"⚠️ Storage persist: {r.status_code} {r.text[:120]}")
    except Exception as e:
        print(f"⚠️ Storage persist error: {e}")

def _get_display_settings() -> dict:
    """Return current settings — from cache or Storage on cold start."""
    global _ds_cache
    if not _ds_cache:
        _ds_cache = _load_from_storage()
    return _ds_cache.copy()

def _save_display_settings(value: dict):
    """Update in-memory cache (instant) and persist to Storage (best-effort)."""
    global _ds_cache
    _ds_cache = value.copy()
    _persist_to_storage(value)

@app.get("/settings")
def get_settings(user=Depends(get_user)):
    return _get_display_settings()

@app.put("/settings")
def update_settings(data: dict, admin=Depends(admin_only)):
    allowed = set(DEFAULT_DISPLAY.keys())
    value = {**DEFAULT_DISPLAY, **{k: bool(v) for k, v in data.items() if k in allowed}}
    _save_display_settings(value)          # always succeeds (in-memory)
    return {"status": "saved", "settings": value}

# ── HTML Daily Report ─────────────────────────────────────────────────────────

@app.get("/report/{rid}", response_class=HTMLResponse)
def html_report(rid: int):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?id=eq.{rid}", headers=sb_headers())
    rows = r.json()
    if not rows: return HTMLResponse("<h2>Report not found</h2>", status_code=404)
    rep = _parse(rows[0])
    ds = _get_display_settings()

    # Conditional info-grid cells
    gps_cell = f'<div class="f"><div class="l">🌐 GPS</div><div class="v">{rep["gps"] or "-"}</div></div>' if ds.get("showGps") else ""
    wt_cell  = f'<div class="f full"><div class="l">🔧 Work Type</div><div class="v">{", ".join(rep["workTypes"]) or "-"}</div></div>' if ds.get("showWorkTypes") else ""
    qty_cell = f'<div class="f"><div class="l">📏 Quantity</div><div class="v">{rep["quantity"] or "-"}</div></div>' if ds.get("showQuantity") else ""
    iss_cell = f'<div class="f"><div class="l">⚠️ Issues</div><div class="v">{rep["issues"] or "-"}</div></div>' if ds.get("showIssues") else ""

    # Conditional sections
    safety_s  = _safety_section(rep.get("safety", {}), show_labor=ds.get("showLabor", True)) if ds.get("showSafety") else ""
    team_s    = _grid("👷 Working Team (ถ่ายรูปกองงานเรียงแถว)", rep["teamImages"]) if ds.get("showTeamImages") else ""
    eq_s      = _grid("🔧 Tools &amp; Machines (เครื่องมือเครื่องจักรที่ใช้)", rep["equipmentImages"]) if ds.get("showEquipmentImages") else ""
    mat_s     = _grid_cap("📦 Materials (วัสดุเข้า-ออก)", rep["materialImages"]) if ds.get("showMaterialImages") else ""
    area_s    = _grid_cap("📍 Work Area (ถ่ายรูปพื้นที่ทำงาน ก่อนและหลัง)", rep["areaImages"]) if ds.get("showAreaImages") else ""
    closing_s = _grid("🔒 ปิดกั้นและคลุมหลุมขุด (ก่อนออกจากสถานที่)", rep.get("closingImages", [])) if ds.get("showClosingImages") else ""

    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>LXT Report – {rep['date']}</title><style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;background:#f0f4f8;padding:20px;color:#111}}
.wrap{{max-width:860px;margin:auto;background:white;border-radius:10px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.12)}}
.hdr{{background:linear-gradient(135deg,#1e3a5f,#1d4ed8);color:white;padding:22px 28px}}
.hdr h1{{font-size:20px;font-weight:800}}.hdr p{{font-size:13px;opacity:.85;margin-top:4px}}
.body{{padding:24px 28px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:20px}}
.f{{background:#f8fafc;border-radius:6px;padding:10px 14px}}
.f .l{{font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.5px}}
.f .v{{font-weight:700;color:#1e3a5f;margin-top:3px;font-size:14px}}
.full{{grid-column:1/-1}}
.section{{margin-bottom:24px}}
.sec-title{{font-size:15px;font-weight:700;color:#1e3a5f;border-bottom:2px solid #1d4ed8;padding-bottom:6px;margin-bottom:12px}}
.photos{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px}}
.photos img{{width:100%;height:180px;object-fit:cover;border-radius:6px;border:1px solid #e5e7eb}}
.photos-cap{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px}}
.photo-item{{display:flex;flex-direction:column}}
.photo-item img{{width:100%;height:180px;object-fit:cover;border-radius:6px;border:1px solid #e5e7eb}}
.photo-caption{{font-size:11px;color:#374151;margin:4px 0 0;text-align:center;word-break:break-word;line-height:1.4;font-style:italic}}
.empty{{color:#9ca3af;font-style:italic;font-size:13px}}
.btn{{background:#1d4ed8;color:white;border:none;padding:10px 24px;border-radius:6px;cursor:pointer;font-size:14px;font-weight:700;margin-bottom:20px}}
.foot{{text-align:center;padding:14px;background:#f8fafc;color:#9ca3af;font-size:12px}}
@media print{{body{{background:white;padding:0}}.wrap{{box-shadow:none}}.btn{{display:none}}.photos img{{break-inside:avoid}}}}
</style></head><body><div class="wrap">
<div class="hdr"><h1>📋 LXT Daily Progress Report</h1><p>LXT Network Co., Ltd.</p></div>
<div class="body">
<button class="btn" onclick="window.print()">🖨️ Print / Save as PDF</button>
<div class="grid">
  <div class="f"><div class="l">📅 Date</div><div class="v">{rep['date']}</div></div>
  <div class="f"><div class="l">🏗️ Project</div><div class="v">{rep['project'] or '-'}</div></div>
  <div class="f"><div class="l">📍 Site</div><div class="v">{rep['site'] or '-'}</div></div>
  {gps_cell}
  {wt_cell}
  <div class="f full"><div class="l">📝 Description</div><div class="v">{rep['description'] or '-'}</div></div>
  {qty_cell}
  {iss_cell}
</div>
{safety_s}{team_s}{eq_s}{mat_s}{area_s}{closing_s}
</div>
<div class="foot">Generated by LXT Daily Report System — {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
</div></body></html>"""
    return HTMLResponse(html)

# ── Excel Export ──────────────────────────────────────────────────────────────

@app.get("/export")
def export_excel(user=Depends(get_user_from_query)):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/reports?order=date.asc,id.asc", headers=sb_headers())
    rows = r.json()
    wb = Workbook(); ws = wb.active; ws.title = "LXT Daily Report"
    hdrs = ["Date","Project","Site","GPS","Work Type","Description","Quantity","Issues",
            "Toolbox Talk","Traffic Signs","Shoring","PPE","First Aid","Labor (คน)","Area Safety",
            "Team Photos","Equipment Photos","Material Photos","Area Photos","Closing Photos"]
    fill_blue   = PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
    fill_green  = PatternFill(start_color="1A5E3A",end_color="1A5E3A",fill_type="solid")
    font_w = Font(color="FFFFFF",bold=True,size=11)
    ws.append(hdrs)
    for i,c in enumerate(ws[1],1):
        c.font=font_w; c.alignment=Alignment(horizontal="center",vertical="center")
        c.fill = fill_green if i in range(9,16) else fill_blue
    for i,w in enumerate([12,25,15,20,25,40,15,30,12,12,12,12,12,10,12,40,40,40,40,40],1):
        ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width=w
    ws.row_dimensions[1].height=22
    for row in rows:
        sf = row.get("safety") or {}
        if isinstance(sf, str):
            try: sf=json.loads(sf)
            except: sf={}
        def chk(k): return "✅" if sf.get(k) else "❌"
        ws.append([row.get("date",""),row.get("project",""),row.get("site",""),row.get("gps",""),
                   ", ".join(json.loads(row.get("work_types") or "[]")),row.get("description",""),
                   row.get("quantity",""),row.get("issues",""),
                   chk("toolboxTalk"),chk("trafficSigns"),chk("shoring"),chk("ppe"),chk("firstAid"),
                   sf.get("laborCount",""),chk("areaSafety"),
                   " | ".join(json.loads(row.get("team_images") or "[]")),
                   " | ".join(json.loads(row.get("equipment_images") or "[]")),
                   " | ".join(i.get("url","") if isinstance(i,dict) else i for i in json.loads(row.get("material_images") or "[]")),
                   " | ".join(i.get("url","") if isinstance(i,dict) else i for i in json.loads(row.get("area_images") or "[]")),
                   " | ".join(json.loads(row.get("closing_images") or "[]"))])
    ws.freeze_panes="A2"
    fname=f"LXT_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fp=f"/tmp/{fname}"; wb.save(fp)
    return FileResponse(path=fp,filename=fname,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
